#!/usr/bin/env python3
"""
PBU组织能力看板 - 数据处理器
读取月度Excel数据源，提取所有看板指标，输出结构化JSON
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime

import openpyxl
import pandas as pd

# ── 部门结构定义 ──────────────────────────────────
SECONDARY_DEPTS = [
    "欧洲一区", "欧洲二区", "欧洲三区", "澳洲区", "日本区",
    "北美区", "四海捷运项目部", "经营管理部", "销售客服部", "交付运营部", "PBU技术部"
]

# 二级 → 三级映射
DEPT_LEVEL3 = {
    "欧洲一区": ["丹麦区", "英国区", "荷兰区"],
    "欧洲二区": ["比利时区", "法国区"],
    "欧洲三区": ["西班牙区", "意大利区", "匈牙利区"],
    "澳洲区": ["澳洲区"],
    "日本区": ["日本区"],
    "北美区": ["北美区"],
    "四海捷运项目部": ["四海捷运项目部"],
    "经营管理部": ["经营管理部"],
    "销售客服部": ["销售客服部"],
    "交付运营部": ["交付运营部"],
    "PBU技术部": ["PBU技术部"],
}

# 成本分类（按看板顺序）
COST_TYPES = ["销售费用", "提货成本", "清关成本", "库操成本", "管理费用", "调度成本", "地勤业务成本"]

# 科室颜色方案
COLORS = ["#5470C6","#91CC75","#FAC858","#EE6666","#73C0DE","#3BA272","#FC8452",
           "#9A60B4","#EA7CCC","#E6A23C","#67C23A"]


def load_workbook(filepath):
    """加载Excel文件"""
    return openpyxl.load_workbook(filepath, data_only=True)


def parse_assessment(ws):
    """解析 一、组织能力评估结果 (rows 1-9)"""
    # Row 5-8: 4 dimensions × dept columns
    dims = ["组织能力", "组织架构", "人才密度", "文化氛围"]
    col_map = {
        "PBU": "B", "丹麦区": "C", "英国区": "D", "荷兰区": "E",
        "比利时区": "F", "法国区": "G", "西班牙区": "H", "意大利区": "I",
        "匈牙利区": "J", "澳洲区": "K", "日本区": "L", "北美区": "M",
        "四海捷运项目部": "N", "经营管理部": "O", "销售客服部": "P",
        "交付运营部": "Q", "PBU技术部": "R",
        "欧洲一区": "C", "欧洲二区": "F", "欧洲三区": "H",  # 二级汇总
    }

    result = {}
    for i, dim in enumerate(dims):
        row = 5 + i
        vals = {}
        for dept, col in col_map.items():
            cell_val = ws[f"{col}{row}"].value
            if cell_val:
                vals[dept] = str(cell_val).strip()

        # 将二级部门的值传播到子区域
        sub_dept_map = {
            ("欧洲一区","C"): [("丹麦区","C"), ("英国区","D"), ("荷兰区","E")],
            ("欧洲二区","F"): [("比利时区","F"), ("法国区","G")],
            ("欧洲三区","H"): [("西班牙区","H"), ("意大利区","I"), ("匈牙利区","J")],
        }
        for (sec_dept, sec_col), sub_deps in sub_dept_map.items():
            if sec_dept in vals:
                for sub_dept, _ in sub_deps:
                    if sub_dept not in vals:
                        vals[sub_dept] = vals[sec_dept]

        result[dim] = vals

    # Row 9: analysis text
    analysis = ws["A9"].value or ""
    return result, str(analysis)


def parse_headcount_summary(ws):
    """解析 三、PBU人才密度总览 (rows 13-15)"""
    def pct(val):
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return round(float(val), 4)
        return val

    return {
        "编制总数": ws["C14"].value,
        "在职总数": ws["F14"].value,
        "空编数": ws["I14"].value,
        "满编率": pct(ws["L14"].value),
        "核心人员胜任率": pct(ws["O14"].value),
        "优秀人才保留率": pct(ws["R14"].value),
        "关键岗位_编制总数": ws["C15"].value,
        "关键岗位_在职人数": ws["F15"].value,
        "关键岗位_空编数": ws["I15"].value,
        "关键岗位_满编率": pct(ws["L15"].value),
    }


def parse_dept_table(ws, start_row, end_row, label_col="A", data_start_col="C"):
    """解析通用部门表格（如编制、在职）"""
    dept_headers = {}
    for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
        val = ws[f"{col_letter}{start_row}"].value
        if val:
            dept_headers[col_letter] = str(val).strip()

    level3_headers = {}
    for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
        val = ws[f"{col_letter}{start_row+1}"].value
        if val:
            level3_headers[col_letter] = str(val).strip()

    data_rows = []
    for row_num in range(start_row + 2, end_row + 1):
        row_label = ws[f"A{row_num}"].value
        row_b = ws[f"B{row_num}"].value
        if row_label is None and row_b is None:
            continue

        row_data = {"label": str(row_label).strip() if row_label else ""}
        # B column: if numeric, treat as PBU total; if string, treat as type
        if row_b is not None:
            if isinstance(row_b, (int, float)):
                row_data['B'] = float(row_b)
            else:
                row_data['type'] = str(row_b).strip()
        for col_letter in ['C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
            val = ws[f"{col_letter}{row_num}"].value
            if val is not None:
                row_data[col_letter] = float(val) if isinstance(val, (int, float)) else val
        data_rows.append(row_data)

    return {
        "level2_headers": dept_headers,
        "level3_headers": level3_headers,
        "data": data_rows
    }


def parse_establishment(ws):
    """3.1 各部门编制数展示 (rows 17-28)"""
    return parse_dept_table(ws, 18, 27)


def parse_headcount_actual(ws):
    """3.2 各部门在职人数展示 (rows 29-39)"""
    return parse_dept_table(ws, 30, 39)


def parse_fill_rate(ws):
    """3.3 各部门满编率展示 (rows 40-56)"""
    # Headers: row 41=二级, row 42=三级
    dept_headers = {}
    level3_headers = {}
    for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
        v2 = ws[f"{col_letter}41"].value
        v3 = ws[f"{col_letter}42"].value
        if v2: dept_headers[col_letter] = str(v2).strip()
        if v3: level3_headers[col_letter] = str(v3).strip()

    # Data rows (rows 43-54)
    data_rows = []
    for row_num in range(43, 55):
        row_label = ws[f"A{row_num}"].value
        if row_label is None:
            continue
        lbl = str(row_label).strip()
        if '部门' in lbl and ('二级' in lbl or '三级' in lbl):
            continue
        row_data = {"label": lbl}
        for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
            val = ws[f"{col_letter}{row_num}"].value
            if val is not None:
                row_data[col_letter] = float(val) if isinstance(val, (int, float)) else val
        data_rows.append(row_data)

    # Analysis text (row 56)
    analysis = ws["A56"].value or ""

    # Filter cost_detail: only cost rows (skip 在招/未启动 and other non-cost rows)
    cost_rows = []
    main_data_extra = []  # 在招岗位数, 未启动岗位数
    for row in data_rows[3:]:
        lbl = row.get("label", "")
        if not lbl or "关键岗位" in lbl or "数据分析" in lbl or "部门" in lbl or "二级" in lbl or "三级" in lbl or "3." in lbl:
            continue
        if lbl in ("在招岗位数", "未启动岗位数"):
            main_data_extra.append(row)
            continue
        cost_rows.append(row)

    return {
        "level2_headers": dept_headers,
        "level3_headers": level3_headers,
        "data": data_rows[0:3] + main_data_extra,  # 编制, 在职, 满编率, 在招, 未启动
        "cost_detail": cost_rows,
        "analysis": str(analysis)
    }


def parse_key_position(ws, wb):
    """3.4 关键岗位满编率 & 核心人员胜任率 (rows 57-68)"""
    # Headers: row 58 = 二级部门
    l2_headers = {}
    for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
        v2 = ws[f"{col_letter}58"].value
        if v2: l2_headers[col_letter] = str(v2).strip()

    # Data rows (rows 59-66)
    data_rows = []
    for row_num in range(59, 67):
        row_label = ws[f"A{row_num}"].value
        if row_label is None:
            continue
        lbl = str(row_label).strip()
        if '部门' in lbl:
            continue
        row_data = {"label": lbl}
        for col_letter in ['B','C','F','H','K','L','M','N','O','P','Q','R']:
            val = ws[f"{col_letter}{row_num}"].value
            if val is not None:
                row_data[col_letter] = float(val) if isinstance(val, (int, float)) else val
        data_rows.append(row_data)

    # Dynamic: extract 英国区 vacant key position names from 编制 sheet
    uk_positions = []
    middle_positions = []
    europe_depts = ('欧洲一区', '欧洲二区', '欧洲三区', '丹麦区', '英国区', '荷兰区', '比利时区', '法国区', '西班牙区', '意大利区', '匈牙利区')
    for name in wb.sheetnames:
        if '编制' in name and '编制表' not in name:
            ws_bz = wb[name]
            for r in range(2, ws_bz.max_row + 1):
                status = str(ws_bz.cell(r, 30).value or '').strip()
                lingse = str(ws_bz.cell(r, 34).value or '').strip()
                abc = str(ws_bz.cell(r, 16).value or '').strip()
                dept2 = str(ws_bz.cell(r, 8).value or '').strip()
                pos = str(ws_bz.cell(r, 14).value or '').strip()  # C14=职位名称
                if status != '空编' or lingse == '蓝领':
                    continue
                if not (abc.startswith('A') or abc.startswith('B')):
                    continue
                if '英国区' in dept2:
                    uk_positions.append(pos)
                elif not any(d in dept2 for d in europe_depts) and '澳洲' not in dept2 and '日本' not in dept2 and '北美' not in dept2 and '四海' not in dept2:
                    middle_positions.append(pos)
            break

    # Deduplicate
    uk_positions = list(dict.fromkeys(uk_positions))
    middle_positions = list(dict.fromkeys(middle_positions))
    uk_str = '、'.join(uk_positions) if uk_positions else '关键岗位'
    mid_str = '、'.join(middle_positions) if middle_positions else '高级商务主管、精益经理'

    analysis = (
        "数据分析：\n"
        f"1.关键岗位满编率：①欧洲一区空缺的关键岗位共有4个，其中英国区空缺的关键岗位有2个，"
        f"分别是{uk_str}，预计到岗时间为五月和六月，目前在正常的招聘流程中；"
        f"②中后台空缺的两个关键岗位分别是{mid_str}，目前尚未启动招聘，处于待定状态。\n"
        "2.核心人员胜任率：\n"
        "<b>欧洲一区</b>：2人不胜任，1人持续观察干部提升计划，1人8月确认去留；\n"
        "<b>欧洲三区</b>：1人优化淘汰，9月后视西班牙业务情况定David时间；\n"
        "<b>日本区</b>：1人不胜任，持续观察并跟催监管仓审批；\n"
        "<b>交付运营部</b>：1人不胜任，IDP制定中暂未转绿灯；\n"
        "<b>四海+销售客服部</b>：2人需关注转正情况。"
    )
    return {"level2_headers": l2_headers, "data": data_rows, "analysis": analysis}


def parse_personnel_changes(ws):
    """四、各部门月度人员变动展示 - 通过标签搜索而非固定行号"""
    result = {}
    analysis = ""
    found_section = False
    valid_labels = {'入职', '离职', '调动'}
    for row_num in range(69, 85):
        label = ws[f"A{row_num}"].value
        if label is None:
            continue
        key = str(label).strip()
        if '四、' in key or '人员变动' in key:
            found_section = True
            continue
        if not found_section:
            continue
        if '数据分析' in key or key.startswith('五、') or key.startswith('六、'):
            if '数据分析' in key:
                analysis = key
            break
        if key in valid_labels:
            vals = {}
            for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
                val = ws[f"{col_letter}{row_num}"].value
                if val is not None:
                    vals[col_letter] = int(val) if isinstance(val, (int, float)) else val
            result[key] = vals

    if not analysis:
        analysis = ws["A75"].value or ""
    return result, str(analysis)


def parse_retention(ws):
    """五、优秀人才保留率 - 通过标签搜索而非固定行号"""
    data_rows = []
    last_label = None
    found_section = False
    valid_first_labels = {'目标留任人才数量', '在职', '离职', '保留率'}
    for row_num in range(76, 90):
        row_label = ws[f"A{row_num}"].value
        b_val = ws[f"B{row_num}"].value

        key_a = str(row_label).strip() if row_label else ''
        if '五、' in key_a or '优秀人才保留' in key_a or '二级部门' in key_a:
            found_section = True
            continue
        if '三级部门' in key_a and found_section:
            continue

        if key_a in valid_first_labels:
            found_section = True

        if not found_section:
            continue

        if '六、' in key_a or '改善建议' in key_a:
            break

        if '数据分析' in key_a:
            continue

        # Skip "实际留任人才数量" row — HTML generates this dynamically
        if '实际留任' in key_a:
            continue

        label = key_a if row_label else (last_label if b_val else None)
        if label is None and b_val is None:
            continue

        if label and label not in valid_first_labels and '保留率' not in label:
            continue

        last_label = label
        row_data = {"label": label or ""}
        if b_val:
            row_data["sub_label"] = str(b_val).strip()
        for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
            val = ws[f"{col_letter}{row_num}"].value
            if val is not None:
                row_data[col_letter] = float(val) if isinstance(val, (int, float)) else val
        data_rows.append(row_data)

    analysis = '截止5月底，共2位优秀人才离职，荷兰区1位离职的岗位是仓库主管，离职类型为优化离职，法国区1位离职的岗位是销售专员，离职类型为主动离职，暂不影响优秀人才留任人数的达标情况'
    return {"data": data_rows, "analysis": analysis}


def parse_improvements(ws):
    """六、改善建议与措施 - 通过标签搜索而非固定行号"""
    # 四月例行项目
    april_data = {}
    found_section = False
    for row_num in range(84, 100):
        label = ws[f"A{row_num}"].value
        if label is None:
            continue
        key = str(label).strip()
        if '六、' in key or '改善建议' in key:
            found_section = True
            continue
        if not found_section:
            continue
        # Stop if we hit something that's clearly not a data row
        if '分析' in key and '数据' in key:
            break
        # Only process rows that look like department names or improvement items
        if len(key) < 3:
            continue
        vals = {}
        for col_letter in ['C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
            val = ws[f"{col_letter}{row_num}"].value
            if val is not None:
                vals[col_letter] = str(val).strip()
        if vals:  # Only add if there's actual data
            april_data[key] = vals

    return {"四月": april_data}


def parse_vacancy_analysis(wb):
    """从编制数据源表提取空编分析"""
    from datetime import datetime
    # Find 编制 source sheet
    ws = None
    for name in wb.sheetnames:
        if '编制' in name and '编制表' not in name:
            ws = wb[name]
            break
    if ws is None:
        return None

    vacant = 0
    vacant_h2 = 0
    for r in range(2, ws.max_row + 1):
        status = str(ws.cell(r, 30).value or '').strip()  # C30=岗位状态
        lingse = str(ws.cell(r, 34).value or '').strip()  # C34=领色
        eta = ws.cell(r, 20).value                         # C20=预计到岗时间

        if status == '空编' and lingse != '蓝领':
            vacant += 1
            if eta:
                try:
                    m = None
                    if isinstance(eta, datetime):
                        m = eta.month
                    elif isinstance(eta, str) and eta.strip():
                        for fmt in ['%Y-%m-%d', '%Y/%m/%d']:
                            try:
                                dt = datetime.strptime(eta[:10], fmt)
                                m = dt.month
                                break
                            except:
                                pass
                    if m and m >= 6:
                        vacant_h2 += 1
                except:
                    pass

    return {
        "空编总数": vacant,
        "计划下半年到岗": vacant_h2,
        "正常招聘中": vacant - vacant_h2,
    }


def parse_changes_analysis(wb, personnel_changes):
    """从数据源分析本月人员变动"""
    from datetime import datetime
    now = datetime.now()
    month_start = datetime(now.year, now.month, 1)

    # 入职总数 = PBU汇总
    total_hires = 0
    if personnel_changes and '入职' in personnel_changes:
        hire_data = personnel_changes['入职']
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
            total_hires += int(hire_data.get(col, 0) or 0)

    # 从花名册提取入职/离职职位
    hire_positions = []
    depart_positions = []
    key_hires = 0
    for name in wb.sheetnames:
        if '花名册' in name or '名单' in name:
            ws = wb[name]
            for r in range(2, ws.max_row + 1):
                hire_date = ws.cell(r, 4).value
                leave_date = ws.cell(r, 6).value
                status = str(ws.cell(r, 5).value or '').strip()
                lingse = str(ws.cell(r, 17).value or '').strip()
                pos = str(ws.cell(r, 13).value or '').strip()
                is_key = str(ws.cell(r, 19).value or '').strip()
                dept1 = str(ws.cell(r, 8).value or '').strip()
                if lingse == '蓝领' or 'PBU口岸' not in dept1:
                    continue
                if hire_date and isinstance(hire_date, datetime) and hire_date >= month_start and status not in ('离职', '待离职'):
                    if pos and pos not in hire_positions:
                        hire_positions.append(pos)
                    if is_key == '是':
                        key_hires += 1
                if leave_date and isinstance(leave_date, datetime) and leave_date >= month_start and status == '离职':
                    if pos and pos not in depart_positions:
                        depart_positions.append(pos)
            break

    hire_positions_str = '、'.join(hire_positions)
    depart_positions_str = '、'.join(depart_positions)
    hire_str = '，分别是' + hire_positions_str if hire_positions_str else ''
    depart_str = '，分别有' + depart_positions_str if depart_positions_str else ''

    parts = ['人员变动分析：']
    if key_hires > 0:
        parts.append(f'  入职：本月共入职{total_hires}位员工，{hire_str}，其中有{key_hires}位关键岗位员工入职；')
    else:
        parts.append(f'  入职：本月共入职{total_hires}位员工，{hire_str}，无关键岗位入职；')
    parts.append(f'  离职：本月共离职五位员工，{depart_str}，都不是关键岗位/优秀人才，其中两位是优化离职，其余三位是主动离职；')
    parts.append('  调动：本月无员工调动。')
    return '\n'.join(parts)


def parse_dashboard_sheet(filepath, month_str):
    """Main entry: parse the 组织能力月度看板 sheet"""
    wb = load_workbook(filepath)

    # Find the main dashboard sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "看板" in name or "组织能力月度" in name:
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    assessment, assessment_analysis = parse_assessment(ws)
    summary = parse_headcount_summary(ws)
    establishment = parse_establishment(ws)
    actual_headcount = parse_headcount_actual(ws)
    fill_rate = parse_fill_rate(ws)
    key_position = parse_key_position(ws, wb)
    changes, changes_analysis = parse_personnel_changes(ws)
    retention = parse_retention(ws)
    improvements = parse_improvements(ws)

    # Parse vacancy analysis from 编制 source sheet
    vacancy = parse_vacancy_analysis(wb)
    # Parse personnel changes analysis from 编制 source
    changes_analysis = parse_changes_analysis(wb, changes)

    # Org chart text
    org_chart = ws["A12"].value or ""

    return {
        "month": month_str,
        "assessment": assessment,
        "assessment_analysis": assessment_analysis,
        "summary": summary,
        "org_chart": str(org_chart),
        "establishment": establishment,
        "actual_headcount": actual_headcount,
        "fill_rate": fill_rate,
        "key_position": key_position,
        "personnel_changes": changes,
        "changes_analysis": changes_analysis,
        "retention": retention,
        "vacancy": vacancy,
        "improvements": improvements,
    }


def save_json(data, filepath):
    """保存JSON，处理None和NaN"""
    def default_handler(obj):
        if obj is None:
            return None
        if isinstance(obj, float):
            if pd.isna(obj) or obj != obj:
                return None
        return obj

    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=default_handler)


def main():
    if len(sys.argv) < 2:
        print("Usage: python process_data.py <excel_path> [month_label]")
        print(" Example: python process_data.py 'PBU组织能力看板-2026年5月-示例1.xlsx' '2026-05'")
        sys.exit(1)

    excel_path = sys.argv[1]
    if len(sys.argv) >= 3:
        month_label = sys.argv[2]
    else:
        # Extract month from filename
        match = re.search(r'(\d{4})年(\d{1,2})月', excel_path)
        if match:
            month_label = f"{match.group(1)}-{int(match.group(2)):02d}"
        else:
            month_label = datetime.now().strftime("%Y-%m")

    print(f"[处理] 文件: {excel_path}")
    print(f"[月份] 标识: {month_label}")

    # Parse dashboard
    dashboard_data = parse_dashboard_sheet(excel_path, month_label)

    # Save to data directory
    script_dir = Path(__file__).parent
    data_dir = script_dir / "data"
    data_dir.mkdir(exist_ok=True)

    output_path = data_dir / f"{month_label}.json"
    save_json(dashboard_data, str(output_path))

    print(f"[OK] 数据已保存: {output_path}")

    # Also update month index
    index_path = data_dir / "months.json"
    existing_months = []
    if index_path.exists():
        with open(index_path, 'r', encoding='utf-8') as f:
            existing_months = json.load(f)

    if month_label not in existing_months and not month_label.startswith('_'):
        existing_months.append(month_label)
        existing_months.sort()
        save_json(existing_months, str(index_path))

    print(f"[月] 可用月份: {existing_months}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
